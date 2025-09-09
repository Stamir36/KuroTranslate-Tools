import openpyxl
from openpyxl.utils import get_column_letter
from lxml import etree
import sys
import os
import hashlib

def create_xliff_from_xlsx_fast(xlsx_path):
    print(f"Открываю файл: {os.path.basename(xlsx_path)}...")
    try:
        # data_only=True для чтения значений, а не формул
        workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"\nОшибка при чтении файла XLSX: {e}")
        input("Нажмите Enter для выхода.")
        return

    # Настройки XLIFF
    XLIFF_NS = "urn:oasis:names:tc:xliff:document:1.2"
    root = etree.Element("xliff", version="1.2", nsmap={None: XLIFF_NS})
    
    file_element = etree.SubElement(root, "file", 
                                    original=os.path.basename(xlsx_path),
                                    source_language="en", 
                                    target_language="ru",
                                    datatype="plaintext")
    body_element = etree.SubElement(file_element, "body")
    
    strings_found = 0
    print("Начинаю высокоскоростной анализ листов...")

    # --- СВЕРХБЫСТРЫЙ МЕТОД ---
    for sheet in workbook.worksheets:
        print(f"-> Анализ листа: {sheet.title}...")
        # 1. Считываем ВСЕ данные с листа в память. Это одна быстрая операция.
        # list() материализует генератор sheet.values в список списков
        data = list(sheet.values)
        
        # 2. Проходим по данным в памяти. Это очень быстро.
        for row_idx, row_data in enumerate(data):
            # Останавливаемся на предпоследней строке, т.к. нам нужна следующая (row_idx + 1)
            if row_idx >= len(data) - 1:
                break
            
            # Ищем все ячейки с 'dialog' в текущей строке
            for col_idx, cell_value in enumerate(row_data):
                if isinstance(cell_value, str) and cell_value.lower() == 'dialog':
                    # Нашли 'dialog'. Берем текст из строки ниже в той же колонке.
                    text_to_translate = data[row_idx + 1][col_idx]
                    
                    if isinstance(text_to_translate, str) and text_to_translate.strip():
                        strings_found += 1
                        trans_id = hashlib.sha1(text_to_translate.encode('utf-8')).hexdigest()[:10]
                        
                        # Вычисляем адрес ячейки (индексы 0-based, адреса 1-based)
                        col_letter = get_column_letter(col_idx + 1)
                        # +2 потому что: +1 для перехода от 0-based к 1-based, и еще +1 т.к. нам нужна следующая строка
                        row_number = row_idx + 2 
                        cell_address = f"{sheet.title}!{col_letter}{row_number}"
                        
                        # Создаем элемент XLIFF
                        trans_unit = etree.SubElement(body_element, "trans-unit", 
                                                      id=trans_id,
                                                      resname=cell_address)
                        source = etree.SubElement(trans_unit, "source")
                        source.text = text_to_translate
                        target = etree.SubElement(trans_unit, "target")
                        target.text = ""

    # Сохранение XLIFF файла
    if strings_found > 0:
        xliff_path = os.path.splitext(xlsx_path)[0] + '.xliff'
        tree = etree.ElementTree(root)
        tree.write(xliff_path, pretty_print=True, xml_declaration=True, encoding='UTF-8')
        print(f"\nАнализ завершен. Найдено строк: {strings_found}")
        print(f"Файл XLIFF успешно создан: {os.path.basename(xliff_path)}")
    else:
        print("\nАнализ завершен. Строк для перевода не найдено.")

    input("Нажмите Enter для выхода.")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        if file_path.endswith('.xlsx'):
            create_xliff_from_xlsx_fast(file_path)
        else:
            print("Пожалуйста, перетащите файл с расширением .xlsx")
            input("Нажмите Enter для выхода.")
    else:
        print("Пожалуйста, перетащите .xlsx файл на этот скрипт.")
        input("Нажмите Enter для выхода.")