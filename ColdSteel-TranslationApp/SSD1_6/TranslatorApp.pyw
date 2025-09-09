import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
import os
import sys
import shutil
import hashlib
import lxml.etree
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.utils import get_column_letter
import xlwings as xw
import re
import subprocess
import threading

# --- КОНСТАНТЫ И НАСТРОЙКИ ---
XLIFF_NAMESPACE_URI = "urn:oasis:names:tc:xliff:document:1.2"
LXML_NS_MAP = {'xliff': XLIFF_NAMESPACE_URI}
NOTE_FILE_REGEX = re.compile(r"^File:\s*(.+)$", re.IGNORECASE)
EDITOR_SCRIPT_NAME = "xliff_editor_gui.py"
DECOMPILER_EXE_NAME = "SenScriptsDecompiler.exe"
DECOMPILED_FOLDER_NAME = "recompiled_files"
COMPILED_SOURCE_FOLDER_NAME = "recompiled_files"
FINAL_DAT_FOLDER_NAME = "complete_dat"

# --- ОСНОВНЫЕ ФУНКЦИИ (ЛОГИКА) ---

def run_external_process(command, log_callback):
    try:
        creationflags = subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0
        process = subprocess.run(
            command, capture_output=True, text=True, encoding='utf-8', 
            errors='replace', creationflags=creationflags
        )
        if process.stdout: log_callback(process.stdout)
        if process.stderr: log_callback(f"ОШИБКА ПРОЦЕССА: {process.stderr}")
        return process.returncode == 0
    except FileNotFoundError:
        log_callback(f"КРИТИЧЕСКАЯ ОШИБКА: Файл '{command[0]}' не найден!")
        return False
    except Exception as e:
        log_callback(f"КРИТИЧЕСКАЯ ОШИБКА при запуске процесса: {e}")
        return False

def decompile_dats(input_folder, log_callback):
    script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
    decompiler_path = os.path.join(script_dir, DECOMPILER_EXE_NAME)
    output_folder_path = os.path.join(script_dir, DECOMPILED_FOLDER_NAME)
    if not os.path.exists(decompiler_path):
        messagebox.showerror("Ошибка", f"Не найден '{DECOMPILER_EXE_NAME}'.")
        return
    dat_files = [f for f in os.listdir(input_folder) if f.lower().endswith('.dat')]
    if not dat_files:
        messagebox.showinfo("Завершено", "В выбранной папке не найдены .dat файлы.")
        return
    log_callback(f"Найдено {len(dat_files)} .dat файлов. Начинаю декомпиляцию...")
    if not os.path.exists(output_folder_path):
        os.makedirs(output_folder_path)
    success_count = 0
    for filename in dat_files:
        filepath = os.path.join(input_folder, filename)
        log_callback(f"\n--- Декомпиляция: {filename} ---")
        if run_external_process([decompiler_path, filepath], log_callback):
            success_count += 1
    log_callback(f"\n--- ДЕКОМПИЛЯЦИЯ ЗАВЕРШЕНА ---\nУспешно: {success_count} из {len(dat_files)}")
    messagebox.showinfo("Готово", f"Декомпиляция завершена.\nУспешно: {success_count} из {len(dat_files)}.\nФайлы в папке '{DECOMPILED_FOLDER_NAME}'.")

def compile_xlsx(log_callback):
    script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
    decompiler_path = os.path.join(script_dir, DECOMPILER_EXE_NAME)
    input_folder_path = os.path.join(script_dir, DECOMPILED_FOLDER_NAME)
    if not os.path.exists(decompiler_path):
        messagebox.showerror("Ошибка", f"Не найден '{DECOMPILER_EXE_NAME}'.")
        return
    if not os.path.exists(input_folder_path):
        messagebox.showerror("Ошибка", f"Папка '{DECOMPILED_FOLDER_NAME}' не найдена.")
        return
    xlsx_files = [f for f in os.listdir(input_folder_path) if f.lower().endswith('.xlsx')]
    if not xlsx_files:
        messagebox.showinfo("Завершено", "Не найдены .xlsx файлы для сборки.")
        return
    log_callback(f"Найдено {len(xlsx_files)} .xlsx файлов. Начинаю сборку...")
    success_count = 0
    for filename in xlsx_files:
        filepath = os.path.join(input_folder_path, filename)
        log_callback(f"\n--- Сборка: {filename} ---")
        if run_external_process([decompiler_path, filepath], log_callback):
            success_count += 1
    log_callback(f"\n--- СБОРКА ЗАВЕРШЕНА ---\nУспешно скомпилировано: {success_count} из {len(xlsx_files)}")
    log_callback("\n--- Перемещение скомпилированных .dat файлов ---")
    source_dat_folder = os.path.join(script_dir, COMPILED_SOURCE_FOLDER_NAME)
    dest_dat_folder = os.path.join(script_dir, FINAL_DAT_FOLDER_NAME)
    if not os.path.isdir(source_dat_folder):
        messagebox.showwarning("Завершено", f"Сборка завершена, но не удалось найти папку '{COMPILED_SOURCE_FOLDER_NAME}'.")
        return
    if not os.path.exists(dest_dat_folder):
        os.makedirs(dest_dat_folder)
    dat_files_to_move = [f for f in os.listdir(source_dat_folder) if f.lower().endswith('.dat')]
    moved_count = 0
    if dat_files_to_move:
        for filename in dat_files_to_move:
            try:
                shutil.move(os.path.join(source_dat_folder, filename), os.path.join(dest_dat_folder, filename))
                moved_count += 1
            except Exception as e:
                log_callback(f"-> ОШИБКА при перемещении '{filename}': {e}")
    final_message = (f"Сборка завершена.\n\nУспешно скомпилировано: {success_count} из {len(xlsx_files)}\nПеремещено в '{FINAL_DAT_FOLDER_NAME}': {moved_count} файлов.")
    log_callback(f"Перемещено {moved_count} .dat файлов.")
    messagebox.showinfo("Готово", final_message)

# --- ИСПРАВЛЕННАЯ ФУНКЦИЯ ---
def create_master_xliff(log_callback):
    script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
    folder_path = os.path.join(script_dir, DECOMPILED_FOLDER_NAME)
    if not os.path.isdir(folder_path):
        messagebox.showerror("Ошибка", f"Папка '{DECOMPILED_FOLDER_NAME}' не найдена.")
        return

    log_callback(f"Начинаю создание XLIFF из папки: {folder_path}")
    
    xlsx_files = sorted([f for f in os.listdir(folder_path) if f.endswith('.xlsx') and not f.startswith('~$')])
    if not xlsx_files:
        messagebox.showerror("Ошибка", f"В папке '{DECOMPILED_FOLDER_NAME}' не найдены .xlsx файлы.")
        return

    xliff_root = ET.Element("xliff", version="1.2", xmlns=XLIFF_NAMESPACE_URI)
    file_element = ET.SubElement(xliff_root, "file", original="master_translation", source_language="en", target_language="ru", datatype="plaintext")
    body_element = ET.SubElement(file_element, "body")
    
    total_strings_found = 0
    
    for filename in xlsx_files:
        log_callback(f"-> Обрабатываю файл: {filename}")
        file_path = os.path.join(folder_path, filename)
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            for sheet in workbook.worksheets:
                data = list(sheet.values)
                for row_idx, row_data in enumerate(data):
                    if row_idx >= len(data) - 1: break
                    for col_idx, cell_value in enumerate(row_data):
                        if isinstance(cell_value, str) and cell_value.lower() == 'dialog':
                            text_to_translate = data[row_idx + 1][col_idx]
                            if isinstance(text_to_translate, str) and text_to_translate.strip():
                                total_strings_found += 1 # Счётчик увеличивается для каждой новой строки
                                
                                col_letter = get_column_letter(col_idx + 1)
                                row_number = row_idx + 2
                                cell_address = f"{sheet.title}!{col_letter}{row_number}"

                                # --- ИСПРАВЛЕНИЕ: Генерируем ID из текста, адреса, ИМЕНИ ФАЙЛА и уникального СЧЁТЧИКА ---
                                unique_string_for_hash = f"{text_to_translate}|{cell_address}|{filename}|{total_strings_found}"
                                trans_id = hashlib.sha1(unique_string_for_hash.encode('utf-8')).hexdigest()[:12]
                                # --------------------------------------------------------------------------------------

                                trans_unit = ET.SubElement(body_element, "trans-unit", id=trans_id)
                                trans_unit.set("resname", cell_address)
                                
                                note_element = ET.SubElement(trans_unit, "note")
                                note_element.text = f"File: {filename}"
                                
                                source = ET.SubElement(trans_unit, "source")
                                source.text = text_to_translate
                                
                                ET.SubElement(trans_unit, "target")
        except Exception as e:
            log_callback(f"Ошибка при обработке файла {filename}: {e}")
            continue
            
    if total_strings_found > 0:
        xliff_path = os.path.join(script_dir, 'translation.xliff')
        tree = ET.ElementTree(xliff_root)
        ET.indent(tree, space="  ", level=0) 
        tree.write(xliff_path, encoding="utf-8", xml_declaration=True)
        log_callback(f"\nНайдено {total_strings_found} строк.\nУспешно создан файл: {xliff_path}")
        messagebox.showinfo("Готово", f"Создан файл 'translation.xliff' с {total_strings_found} строками.")
    else:
        log_callback("\nСтрок для перевода не найдено.")
        messagebox.showwarning("Завершено", "Строк для перевода в файлах не найдено.")

def apply_translation_from_xliff(xliff_path, log_callback):
    script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
    folder_path = os.path.join(script_dir, DECOMPILED_FOLDER_NAME)
    if not os.path.isdir(folder_path):
        messagebox.showerror("Ошибка", f"Папка '{DECOMPILED_FOLDER_NAME}' не найдена.")
        return
    log_callback("Начинаю операцию: Применение перевода...")
    try:
        parser = lxml.etree.XMLParser(remove_blank_text=True)
        tree = lxml.etree.parse(xliff_path, parser)
        translations = {}
        for trans_unit in tree.xpath('//xliff:trans-unit', namespaces=LXML_NS_MAP):
            target_tag = trans_unit.find('xliff:target', namespaces=LXML_NS_MAP)
            resname = trans_unit.get('resname')
            file_origin = None
            note_node = trans_unit.find('xliff:note', namespaces=LXML_NS_MAP)
            if note_node is not None and note_node.text:
                match = NOTE_FILE_REGEX.match(note_node.text.strip())
                if match: file_origin = match.group(1).strip()
            if target_tag is not None and target_tag.text and file_origin and resname:
                target_text = target_tag.text.strip()
                if target_text:
                    if file_origin not in translations: translations[file_origin] = {}
                    translations[file_origin][resname] = target_text
        if not translations:
            messagebox.showwarning("Внимание", "В XLIFF файле не найдено заполненных переводов.")
            return
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось прочитать XLIFF файл: {e}")
        return
    app = None
    total_files = len(translations)
    processed_files = 0
    log_callback(f"Найдено переводов для {total_files} файлов. Начинаю обновление...")
    try:
        app = xw.App(visible=False)
        app.display_alerts = False
        for filename, changes in sorted(translations.items()):
            processed_files += 1
            log_callback(f"({processed_files}/{total_files}) Обновляю файл: {filename}")
            xlsx_path = os.path.join(folder_path, filename)
            if not os.path.exists(xlsx_path):
                log_callback(f"-> ПРЕДУПРЕЖДЕНИЕ: Файл {filename} не найден. Пропускаю.")
                continue
            shutil.copy(xlsx_path, xlsx_path + '.bak')
            book = app.books.open(xlsx_path)
            updated_count = 0
            for cell_address, translated_text in changes.items():
                try:
                    sheet_name, cell_coord = cell_address.split('!', 1)
                    sheet = book.sheets[sheet_name]
                    sheet.range(cell_coord).value = translated_text
                    updated_count += 1
                except Exception as e:
                    log_callback(f"-> Ошибка при обновлении ячейки {cell_address}: {e}")
            book.save()
            book.close()
            log_callback(f"-> Сохранено. Обновлено {updated_count} строк.")
        messagebox.showinfo("Готово", "Все переводы успешно применены!")
        log_callback("\nОперация успешно завершена.")
    except Exception as e:
        messagebox.showerror("Критическая ошибка", f"Ошибка во время работы с Excel: {e}")
    finally:
        if app: app.quit()

def launch_editor(log_callback):
    script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
    editor_script_path = os.path.join(script_dir, EDITOR_SCRIPT_NAME)
    if not os.path.exists(editor_script_path):
        messagebox.showerror("Ошибка", f"Не найден скрипт редактора '{EDITOR_SCRIPT_NAME}'.")
        return
    command = [sys.executable, editor_script_path]
    xliff_path = os.path.join(script_dir, "translation.xliff")
    if os.path.exists(xliff_path):
        command.append(xliff_path)
    else:
        if not messagebox.askyesno("Файл не найден", "Файл 'translation.xliff' не найден.\nВсе равно запустить редактор?"):
            return
    log_callback(f"Запускаю редактор...")
    subprocess.Popen(command)
    log_callback("Редактор запущен.")

# --- КЛАСС ПРИЛОЖЕНИЯ (GUI) ---
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Инструментарий Переводчика")
        self.geometry("900x550")
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=0)
        self.grid_rowconfigure(0, weight=1)
        self.log_textbox = ctk.CTkTextbox(self, state="disabled", wrap="word", font=("Courier New", 12))
        self.log_textbox.grid(row=0, column=0, padx=(15, 10), pady=15, sticky="nsew")
        self.button_frame = ctk.CTkFrame(self)
        self.button_frame.grid(row=0, column=1, padx=(10, 15), pady=15, sticky="ns")
        button_font = ctk.CTkFont(size=13)
        btn_pady = 8; btn_padx = 15
        ctk.CTkLabel(self.button_frame, text="Рабочий процесс", font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(15, 20), padx=btn_padx)
        self.decompile_button = ctk.CTkButton(self.button_frame, text="1. Декомпилировать DAT", command=self.run_decompile, font=button_font)
        self.decompile_button.pack(fill="x", padx=btn_padx, pady=btn_pady)
        self.create_button = ctk.CTkButton(self.button_frame, text="2. Создать XLIFF", command=self.run_create_xliff, font=button_font)
        self.create_button.pack(fill="x", padx=btn_padx, pady=btn_pady)
        self.launch_editor_button = ctk.CTkButton(self.button_frame, text="3. Запустить редактор", command=self.run_launch_editor, font=button_font)
        self.launch_editor_button.pack(fill="x", padx=btn_padx, pady=btn_pady)
        self.apply_button = ctk.CTkButton(self.button_frame, text="4. Применить перевод", command=self.run_apply_translation, font=button_font)
        self.apply_button.pack(fill="x", padx=btn_padx, pady=btn_pady)
        self.compile_button = ctk.CTkButton(self.button_frame, text="5. Собрать в DAT", command=self.run_compile, font=button_font)
        self.compile_button.pack(fill="x", padx=btn_padx, pady=btn_pady)

    def log(self, message):
        def _log():
            self.log_textbox.configure(state="normal")
            self.log_textbox.insert("end", message.strip() + "\n")
            self.log_textbox.configure(state="disabled")
            self.log_textbox.see("end")
        self.after(0, _log)

    def clear_log(self):
        self.log_textbox.configure(state="normal")
        self.log_textbox.delete("1.0", "end")
        self.log_textbox.configure(state="disabled")

    def run_threaded(self, target_func, *args):
        thread = threading.Thread(target=target_func, args=args, daemon=True)
        thread.start()

    def run_decompile(self):
        self.clear_log()
        folder = filedialog.askdirectory(title="Выберите папку с .dat файлами для декомпиляции")
        if folder:
            self.log(f"Выбрана папка с DAT: {folder}")
            self.run_threaded(decompile_dats, folder, self.log)

    def run_compile(self):
        self.clear_log()
        self.run_threaded(compile_xlsx, self.log)

    def run_create_xliff(self):
        self.clear_log()
        self.run_threaded(create_master_xliff, self.log)

    def run_launch_editor(self):
        self.clear_log()
        launch_editor(self.log)

    def run_apply_translation(self):
        self.clear_log()
        script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
        xliff_file = filedialog.askopenfilename(
            title="Выберите 'translation.xliff' файл", filetypes=[("XLIFF files", "*.xliff")],
            initialdir=script_dir, initialfile='translation.xliff'
        )
        if xliff_file:
            self.log(f"Выбран XLIFF файл: {xliff_file}")
            self.run_threaded(apply_translation_from_xliff, xliff_file, self.log)

if __name__ == "__main__":
    app = App()
    app.mainloop()